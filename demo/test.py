# -*- coding: utf-8 -*-
# @Author       :junjie
# @Time         :2020/7/6 15:26
# @FileName     :test.py
# @Motto        :AS the tree,so the fruit
#IDE            :PyCharm

import requests
from openpyxl import load_workbook


header = {
    "Admin-Token": "ern8am8vuoncvoop0o82od98ngs0nmej",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36"
}
session = requests.Session()
session.headers.update(header)
userName = {
    0: "普通用户",
    1: "会员",
    2: "创客",
    3: "金牌",
    4: "联创",
    5: "总公司",
}

i = 1
wb = load_workbook("test.xlsx")
ws = wb.active  # 默认第一个表单名
for a in range(1, 1085):
    url = f'http://pgw.app.lexuan.cn/lexuan-happy/admin/user/account/list?page={a}&row=100'
    data = session.request("GET", url).json()
    for ll in data['data']['items']:
        ws.cell(i + 1, 1, value=ll['userId'])
        ws.cell(i + 1, 2, value=ll['nickname'])
        ws.cell(i + 1, 3, value=ll['mobile'])
        ws.cell(i + 1, 4, value=userName[ll['level']])
        ws.cell(i + 1, 5, value=ll['teamNum'])
        ws.cell(i + 1, 6, value=ll['totalBuy'])
        ws.cell(i + 1, 7, value=ll['totalConsume'])
        ws.cell(i + 1, 8, value=ll['totalIncome'])
        ws.cell(i + 1, 9, value=ll['totalUavaIncome'])
        ws.cell(i + 1, 10, value=ll['notTxAvaBal'])
        ws.cell(i + 1, 11, value=ll['avaBal'])
        ws.cell(i + 1, 12, value=ll['sctBal'])
        print(ll['userId'],ll['nickname'],userName[ll['level']],ll['teamNum'],ll['totalBuy'],ll['totalConsume'],ll['totalIncome'],ll['totalUavaIncome'],ll['notTxAvaBal'],ll['avaBal'],ll['sctBal'])
        i += 1
    wb.save("test.xlsx")
    print(f"第{a}页数据写入成功~")
